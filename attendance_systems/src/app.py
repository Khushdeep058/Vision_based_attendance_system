import tkinter as tk
from tkinter import StringVar
import threading
import os
from datetime import datetime
import openpyxl
import tensorflow as tf

from src.com_in_ineuron_ai_predictor.facePredictor import FacePredictor
from src.com_in_ineuron_ai_collect_trainingdata.get_faces_from_camera import TrainingDataCollector
from src.com_in_ineuron_ai_face_embedding.faces_embedding import GenerateFaceEmbedding
from src.com_in_ineuron_ai_training.train_softmax import TrainFaceRecogModel

EXCEL_FILE = "Attendance.xlsx"

# Disable eager execution for TF1 compatibility
tf.compat.v1.disable_eager_execution()
global_graph = tf.compat.v1.get_default_graph()
global_sess = tf.compat.v1.Session()
tf.compat.v1.keras.backend.set_session(global_sess)

class RegistrationModule:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Online Attendance System")
        self.window.geometry("880x600")
        self.window.configure(background='#ffffff')
        self.window.resizable(0,0)

        header = tk.Label(self.window, text="Online Attendance System", width=80, height=2, fg="white", bg="#363e75",
                          font=('times', 18, 'bold', 'underline'))
        header.place(x=0, y=0)

        # Labels & Entries
        tk.Label(self.window, text="Enroll No", fg="white", bg="#363e75", font=('times', 15)).place(x=80, y=80)
        self.enrollNoTxt = tk.Entry(self.window, width=20, font=('times',15,'bold'))
        self.enrollNoTxt.place(x=205, y=80)

        tk.Label(self.window, text="Roll Number", fg="white", bg="#363e75", font=('times', 15)).place(x=450, y=80)
        self.rollNumberTxt = tk.Entry(self.window, width=20, font=('times',15,'bold'))
        self.rollNumberTxt.place(x=575, y=80)

        tk.Label(self.window, text="Name", fg="white", bg="#363e75", font=('times',15)).place(x=80, y=140)
        self.studentNameTxt = tk.Entry(self.window, width=20, font=('times',15,'bold'))
        self.studentNameTxt.place(x=205, y=140)

        tk.Label(self.window, text="Email ID", fg="white", bg="#363e75", font=('times',15)).place(x=450, y=140)
        self.emailIDTxt = tk.Entry(self.window, width=20, font=('times',15,'bold'))
        self.emailIDTxt.place(x=575, y=140)

        tk.Label(self.window, text="Mobile No", fg="white", bg="#363e75", font=('times',15)).place(x=80, y=200)
        self.mobileNoTxt = tk.Entry(self.window, width=20, font=('times',15,'bold'))
        self.mobileNoTxt.place(x=205, y=200)

        tk.Label(self.window, text="Notification:", fg="white", bg="#363e75", font=('times',15)).place(x=80, y=260)
        self.message = tk.Label(self.window, text="", bg="#bbc7d4", fg="black", width=58, height=2, font=('times',15))
        self.message.place(x=205, y=260)

        # Buttons
        tk.Button(self.window, text="Take Images", command=self.collectUserImageForRegistration,
                  fg="white", bg="#363e75", width=15, height=2, font=('times',15,'bold')).place(x=80, y=350)

        tk.Button(self.window, text="Train Images", command=self.trainModel,
                  fg="white", bg="#363e75", width=15, height=2, font=('times',15,'bold')).place(x=350, y=350)

        tk.Button(self.window, text="Predict", command=self.startPredictionThread,
                  fg="white", bg="#363e75", width=15, height=2, font=('times',15,'bold')).place(x=600, y=350)

        tk.Button(self.window, text="Quit", command=self.close_window,
                  fg="white", bg="#363e75", width=10, height=2, font=('times',15,'bold')).place(x=650, y=510)

        # Recognized students display
        self.recognizedText = tk.Text(self.window, width=60, height=10, font=('times',12))
        self.recognizedText.place(x=150, y=420)
        self.recognizedText.insert(tk.END, "Recognized students will appear here...\n")
        self.recognizedText.config(state=tk.DISABLED)

        # Ensure Excel exists
        self.initialize_excel()

        # Initialize FacePredictor in main thread
        self.faceDetector = FacePredictor()

        self.window.mainloop()

    # ---------------- Excel Functions ---------------- #
    def initialize_excel(self):
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            ws_students = wb.active
            ws_students.title = "Students"
            ws_students.append(["Enroll No","Roll No","Name","Mobile No"])
            ws_attendance = wb.create_sheet("Attendance")
            ws_attendance.append(["Enroll No","Roll No","Name","Date","Status"])
            wb.save(EXCEL_FILE)
            wb.close()

    def add_student_to_excel(self, enroll_no, roll_no, name, mobile_no):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Students"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == enroll_no:
                wb.close()
                return
        ws.append([enroll_no, roll_no, name, mobile_no])
        wb.save(EXCEL_FILE)
        wb.close()

    def mark_attendance(self, enroll_no, roll_no, name):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Attendance"]
        today = datetime.today().strftime("%Y-%m-%d")
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == enroll_no and row[3].value == today:
                wb.close()
                return
        ws.append([enroll_no, roll_no, name, today, "Present"])
        wb.save(EXCEL_FILE)
        wb.close()

    # ---------------- Image/Model Functions ---------------- #
    def collectUserImageForRegistration(self):
        enrollNoVal = self.enrollNoTxt.get()
        rollNoVal = self.rollNumberTxt.get()
        name = self.studentNameTxt.get()
        mobile_no = self.mobileNoTxt.get()

        # Collect images
        trnngDataCollctrObj = TrainingDataCollector({"faces":50, "output":"../datasets/train/"+name})
        trnngDataCollctrObj.collectImagesFromCamera()

        self.add_student_to_excel(enrollNoVal, rollNoVal, name, mobile_no)
        self.message.configure(text=f"Collected 50 images for {name} (Enroll: {enrollNoVal}).")

    def getFaceEmbedding(self):
        genFaceEmbdng = GenerateFaceEmbedding({"dataset":"../datasets/train", "embeddings":"faceEmbeddingModels/embeddings.pickle"})
        genFaceEmbdng.genFaceEmbedding()

    def trainModel(self):
        self.getFaceEmbedding()
        faceRecogModel = TrainFaceRecogModel({
            "embeddings":"faceEmbeddingModels/embeddings.pickle",
            "model":"faceEmbeddingModels/my_model.h5",
            "le":"faceEmbeddingModels/le.pickle"
        })
        faceRecogModel.trainKerasModelForFaceRecognition()
        self.message.configure(text="Model training complete. Ready for prediction.")

    # ---------------- Prediction ---------------- #
    def startPredictionThread(self):
        threading.Thread(target=self.makePrediction, daemon=True).start()
        self.message.configure(text="Face recognition started...")

    def makePrediction(self):
        import cv2

        # Disable OpenCV async warnings
        cv2.setNumThreads(0)

        recognized = []
        try:
            # Run TF/Keras operations inside global graph/session
            with global_graph.as_default():
                tf.compat.v1.keras.backend.set_session(global_sess)
                recognized = self.faceDetector.detectFace()
        except Exception as e:
            print("Error during face recognition:", e)

        if recognized:
            for name, roll, enroll in recognized:
                self.mark_attendance(enroll, roll, name)
                self.recognizedText.config(state=tk.NORMAL)
                self.recognizedText.insert(tk.END, f"{name} marked present.\n")
                self.recognizedText.see(tk.END)
                self.recognizedText.config(state=tk.DISABLED)
            self.message.configure(text="Face recognition session ended.")
        else:
            self.message.configure(text="No faces recognized.")

    # ---------------- Utility ---------------- #
    def close_window(self):
        self.window.destroy()


if __name__ == "__main__":
    RegistrationModule()

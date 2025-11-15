import os
import cv2
import dlib
import numpy as np
import pickle
import datetime
import openpyxl
from mtcnn import MTCNN
from keras.models import load_model
from src.insightface.deploy import face_model
import face_preprocess
from src.com_in_ineuron_ai_detectfaces_mtcnn.Configurations import ConfigurationsPOJO

EXCEL_FILE = "Attendance.xlsx"

class FacePredictor:
    def __init__(self):
        self.trackers = []
        self.texts = []

        self.image_size = '112,112'
        self.model_path = "./insightface/models/model-y1-test2/model,0"
        self.threshold = 1.24
        self.det = 0

        # Initialize detector & embedding model
        self.detector = MTCNN()
        self.embedding_model = face_model.FaceModel(self.image_size, self.model_path, self.threshold, self.det)#embedding_model ek neural network hai jo face se 128-dimensional vector (embedding) nikalta hai — jaise digital signature of a face.

        # Load embeddings and label encoder
        self.embeddings_file = "./faceEmbeddingModels/embeddings.pickle"
        self.le_file = "./faceEmbeddingModels/le.pickle"
        self.data = pickle.loads(open(self.embeddings_file, "rb").read())
        self.le = pickle.loads(open(self.le_file, "rb").read())
        self.embeddings = np.array(self.data['embeddings'])
        self.labels = np.array(self.le.fit_transform(self.data['names']))

        # Load classifier model
        self.model = load_model(ConfigurationsPOJO.clssfr_ModelPath)

        # Thresholds
        self.cosine_threshold = 0.8
        self.proba_threshold = 0.85
        self.comparing_num = 5

        # Ensure Excel exists
        self.initialize_excel()

    # ---------------- Cosine Similarity ---------------- #
    @staticmethod
    def findCosineDistance(vector1, vector2):
        vec1 = vector1.flatten()
        vec2 = vector2.flatten()
        a = np.dot(vec1.T, vec2)
        b = np.dot(vec1.T, vec1)
        c = np.dot(vec2.T, vec2)
        return 1 - (a / (np.sqrt(b) * np.sqrt(c)))

    def CosineSimilarity(self, test_vec, source_vecs):
        cos_dist = 0
        for source_vec in source_vecs:
            cos_dist += self.findCosineDistance(test_vec, source_vec)
        return cos_dist / len(source_vecs)

    # ---------------- Excel Functions ---------------- #
    def initialize_excel(self):
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            ws_students = wb.active
            ws_students.title = "Students"
            ws_students.append(["Enroll No", "Roll No", "Name", "Mobile No"])
            ws_attendance = wb.create_sheet("Attendance")
            ws_attendance.append(["Enroll No", "Roll No", "Name", "Date", "Status"])
            wb.save(EXCEL_FILE)
            wb.close()

    def mark_attendance_in_excel(self, enroll_no, roll_no, name):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_students = wb["Students"]
        ws_attendance = wb["Attendance"]

        # Add student if not exists
        student_exists = any(row[0].value == enroll_no for row in ws_students.iter_rows(min_row=2))
        if not student_exists:
            ws_students.append([enroll_no, roll_no, name, ""])

        # Mark attendance
        today = datetime.date.today().strftime("%Y-%m-%d")
        attendance_exists = any(row[0].value == enroll_no and row[3].value == today for row in ws_attendance.iter_rows(min_row=2))
        if not attendance_exists:
            ws_attendance.append([enroll_no, roll_no, name, today, "Present"])

        wb.save(EXCEL_FILE)
        wb.close()

    def get_student_info_by_name(self, name):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Students"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == name:
                wb.close()
                return row[0], row[1]
        wb.close()
        return None, None

    # ---------------- Face Detection ---------------- #
    def detectFace(self):
        cap = cv2.VideoCapture(0)
        frame_width = int(cap.get(3))
        frame_height = int(cap.get(4))
        save_width = 800
        save_height = int(800 / frame_width * frame_height)

        frames = 0
        recognized_students = set()
        recognized_list = []

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            frames += 1
            frame = cv2.resize(frame, (save_width, save_height))
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

            # Every 3 frames, detect faces
            if frames % 3 == 0:
                self.trackers = []
                self.texts = []

                bboxes = self.detector.detect_faces(frame)
                for bboxe in bboxes:
                    bbox = bboxe['box']
                    bbox = np.array([bbox[0], bbox[1], bbox[0] + bbox[2], bbox[1] + bbox[3]])
                    landmarks = bboxe['keypoints']
                    landmarks = np.array([
                        landmarks["left_eye"][0], landmarks["right_eye"][0], landmarks["nose"][0],
                        landmarks["mouth_left"][0], landmarks["mouth_right"][0],
                        landmarks["left_eye"][1], landmarks["right_eye"][1], landmarks["nose"][1],
                        landmarks["mouth_left"][1], landmarks["mouth_right"][1]
                    ])
                    landmarks = landmarks.reshape((2, 5)).T
                    nimg = face_preprocess.preprocess(frame, bbox, landmarks, image_size='112,112')
                    nimg = cv2.cvtColor(nimg, cv2.COLOR_BGR2RGB)
                    nimg = np.transpose(nimg, (2, 0, 1))
                    embedding = self.embedding_model.get_feature(nimg).reshape(1, -1)

                    text = "Unknown"
                    preds = self.model.predict(embedding).flatten()
                    j = np.argmax(preds)
                    proba = preds[j]

                    match_class_idx = (self.labels == j)
                    match_class_idx = np.where(match_class_idx)[0]
                    selected_idx = np.random.choice(match_class_idx, self.comparing_num)
                    compare_embeddings = self.embeddings[selected_idx]
                    cos_similarity = self.CosineSimilarity(embedding, compare_embeddings)

                    if cos_similarity < self.cosine_threshold and proba > self.proba_threshold:
                        name = self.le.classes_[j]
                        text = name
                        if name not in recognized_students:
                            enroll_no, roll_no = self.get_student_info_by_name(name)
                            if enroll_no and roll_no:
                                self.mark_attendance_in_excel(enroll_no, roll_no, name)
                                recognized_students.add(name)
                                recognized_list.append((name, roll_no, enroll_no))
                        print(f"Recognized: {name} <{proba*100:.2f}>")

                    tracker = dlib.correlation_tracker()
                    rect = dlib.rectangle(bbox[0], bbox[1], bbox[2], bbox[3])
                    tracker.start_track(rgb, rect)
                    self.trackers.append(tracker)
                    self.texts.append(text)

                    y = bbox[1] - 10 if bbox[1] - 10 > 10 else bbox[1] + 10
                    cv2.putText(frame, text, (bbox[0], y), cv2.FONT_HERSHEY_SIMPLEX, 0.95, (255,255,255),1)
                    cv2.rectangle(frame, (bbox[0], bbox[1]), (bbox[2], bbox[3]), (179,0,149), 4)

            # Track previously detected faces
            else:
                for tracker, text in zip(self.trackers, self.texts):
                    pos = tracker.get_position()
                    startX = int(pos.left())
                    startY = int(pos.top())
                    endX = int(pos.right())
                    endY = int(pos.bottom())
                    cv2.rectangle(frame, (startX, startY), (endX, endY), (179,0,149),4)
                    cv2.putText(frame, text, (startX, startY-15), cv2.FONT_HERSHEY_SIMPLEX, 0.95,(255,255,255),1)

            cv2.imshow("Face Recognition Attendance", frame)
            key = cv2.waitKey(1) & 0xFF
            if key == ord("q"):
                break

        cap.release()
        cv2.destroyAllWindows()
        return recognized_list
